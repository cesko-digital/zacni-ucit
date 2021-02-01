from openpyxl import load_workbook


def import_msmt_college_registry_xlsx(path, sheet_name):
    """
    Import XLSX from https://regvssp.msmt.cz/registrvssp/cvslist.aspx
    (list of colleges and faculties).

    Parameters:
        path -- path to XLSX file
        sheet_name -- "ExportVS" or "ExportFakulty"
    """
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]

    out = []
    columns = [k.value.strip() for k in sheet[1]]
    for i in range(2, sheet.max_row + 1):
        values = [i.value for i in sheet[i]]
        item = dict(zip(columns, values))
        out.append(item)

    return out
